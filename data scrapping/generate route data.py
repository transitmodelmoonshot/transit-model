import os
from inspect import getsourcefile
from os.path import abspath
import numpy as np
import openpyxl #opens excel files
from openpyxl import Workbook
#set active directory to app location
directory = abspath(getsourcefile(lambda:0))
#check if system uses forward or backslashes for writing directories
if(directory.rfind("/") != -1):
    newDirectory = directory[:(directory.rfind("/")+1)]
else:
    newDirectory = directory[:(directory.rfind("\\")+1)]
os.chdir(newDirectory)

def subNone(value):
    if value is None:
        return(0)
    else:
        return(value)

inp = openpyxl.load_workbook("../VRTS data pull.xlsx").active
out_wb = Workbook()
out_wb.title = "Route level data"
hrs = out_wb.active
hrs.title = "Service Hours"
bords = out_wb.create_sheet("Boardings")
for i in range(2,9615):
    year = inp["A{}".format(i)].value
    ISOweek = inp["B{}".format(i)].value
    route = inp["C{}".format(i)].value
    boardings = subNone(inp["H{}".format(i)].value)
    trips = subNone(inp["E{}".format(i)].value)
    runtime = subNone(inp["J{}".format(i)].value)

    hours = trips*runtime/60
    if year == 2019:
        week_counter = ISOweek
    elif year == 2020:
        week_counter = 52 + ISOweek
    elif year == 2021:
        week_counter = 52 + 53 + ISOweek
    elif year == 2022:
        week_counter = 52 + 53 + 52 + ISOweek
    hrs.cell(row=week_counter+1,column = route+1).value = hours
    bords.cell(row=week_counter+1,column = route+1).value = boardings
    print("{} {} {} {} {} {}".format(year, ISOweek,route,trips,hours,week_counter))
out_wb.save("../route level data.xlsx")
