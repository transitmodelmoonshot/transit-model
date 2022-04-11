import os
from inspect import getsourcefile
from os.path import abspath
import numpy as np
import openpyxl #opens excel files
#set active directory to app location
directory = abspath(getsourcefile(lambda:0))
#check if system uses forward or backslashes for writing directories
if(directory.rfind("/") != -1):
    newDirectory = directory[:(directory.rfind("/")+1)]
else:
    newDirectory = directory[:(directory.rfind("\\")+1)]
os.chdir(newDirectory)

input = openpyxl.load_workbook("../deluxe data.xlsx").active
wb = openpyxl.load_workbook("../route level data.xlsx")
output = wb["daily"]
i = 2
j = 1
prev_week = 0
prev_date = None
output.cell(row = 1, column = 1, value = "ISO Week")
while(1):
        route = input["C{}".format(i)].value
        week = input["B{}".format(i)].value
        date = input["A{}".format(i)].value
        if route is None:
            break
        if date != prev_date:
            j = j+1
        x = input["D{}".format(i)].value
        if x is None:
            x = 0

        output.cell(row = j,column = route+2,value = input["D{}".format(i)].value)
        output.cell(row = 1, column = route+2, value = "Route {}".format(route))
        output.cell(row = j, column = 2, value = week)
        output.cell(row = j, column = 1, value = date)
        #print("j is {}, i is {}, route is {}".format(j,i,route))
        i = i+1
        prev_week = week
        prev_date = date

input = wb["daily"]
output = wb["weekly"]
for column in range(3,91):
    vals = []
    output_row = 1
    old_week = 1
    for row in range(2,1141):
        x = input.cell(row = row,column = column).value
        week = input.cell(row = row, column = 2).value
        print("Row {}. x is {} week is {} old week is {}".format(row,x,week,old_week))
        if x is None:
            x = 0
        if old_week != week:
            output_row = output_row+1
            output.cell(row = output_row,column = 1, value = old_week)
            output.cell(row = output_row,column = column,value = np.sum(vals))
            vals = [x]
            old_week = week
        else:
            vals.append(x)
wb.save('../route level data.xlsx')
