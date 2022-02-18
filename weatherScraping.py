import os
from inspect import getsourcefile
from os.path import abspath
import openpyxl #opens excel files
#set active directory to app location
directory = abspath(getsourcefile(lambda:0))
#check if system uses forward or backslashes for writing directories
if(directory.rfind("/") != -1):
    newDirectory = directory[:(directory.rfind("/")+1)]
else:
    newDirectory = directory[:(directory.rfind("\\")+1)]
os.chdir(newDirectory)

path = "yyj climate data.xlsx"
def func():
    wb = openpyxl.load_workbook(path)
    sh_daily = wb["daily"]
    sh_weekly = wb["weekly"]
    day = 1
    week = 1
    cell = sh_daily.cell(row = day, column = 14)
    print(cell.value)
    while(1):
        average_temp = 0
        average_daily_temp = 0
        for i in range(1,8):
            day = day+1
            daily_temp = sh_daily.cell(row = day, column = 14)
            sh_daily["I{}".format(day)] = week
            if(daily_temp.value is None):
                wb.save(path)
                return

            average_daily_temp = average_daily_temp+daily_temp.value
        sh_weekly["A{}".format(week)] = week
        sh_weekly["B{}".format(week)] = average_daily_temp/7
        week = week+1

func()
